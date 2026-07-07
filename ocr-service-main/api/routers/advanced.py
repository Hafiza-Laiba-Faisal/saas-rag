"""Advanced OCR features: searchable PDF, Excel, preprocessing, barcode,
document classification, layout visualization, job queue, metrics."""
from __future__ import annotations

import io
import json
import time
import uuid
import asyncio
from collections import defaultdict
from pathlib import Path
from typing import Any, Optional

import cv2
import numpy as np
from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel

from api.dependencies import get_image_service, get_settings
from config.settings import Settings
from services.image_service import ImageOCRService
from services.validator import FileValidator
from utils.request_id import generate_request_id
from utils.logger import get_logger

logger = get_logger(__name__)
router = APIRouter(prefix="/ocr", tags=["Advanced"])

# ─────────────────────────────────────────────
# In-memory job store
# ─────────────────────────────────────────────
_jobs: dict[str, dict] = {}
_metrics: dict[str, Any] = {
    "total_requests": 0,
    "successful": 0,
    "failed": 0,
    "total_words": 0,
    "total_processing_ms": 0.0,
    "by_endpoint": defaultdict(int),
    "recent_errors": [],
}


def _record(endpoint: str, words: int = 0, ms: float = 0.0, success: bool = True, error: str = ""):
    _metrics["total_requests"] += 1
    if success:
        _metrics["successful"] += 1
        _metrics["total_words"] += words
        _metrics["total_processing_ms"] += ms
    else:
        _metrics["failed"] += 1
        if error:
            _metrics["recent_errors"] = ([{"endpoint": endpoint, "error": error, "ts": time.time()}]
                                          + _metrics["recent_errors"])[:20]
    _metrics["by_endpoint"][endpoint] += 1


# ─────────────────────────────────────────────
# 1. SEARCHABLE PDF EXPORT
# ─────────────────────────────────────────────
@router.post("/export/searchable-pdf")
async def export_searchable_pdf(
    file: UploadFile = File(...),
    image_service: ImageOCRService = Depends(get_image_service),
    settings: Settings = Depends(get_settings),
) -> Response:
    """Run OCR and return a searchable PDF with an invisible text layer."""
    request_id = generate_request_id()
    validator = FileValidator()
    content = await validator.validate_upload(file, settings)

    t0 = time.monotonic()
    result = image_service.process(content, file.filename or "image", request_id)
    ms = (time.monotonic() - t0) * 1000

    try:
        import fitz  # PyMuPDF
        arr = np.frombuffer(content, np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        h, w = img.shape[:2]

        # Encode original image as PNG bytes
        _, png_buf = cv2.imencode(".png", img)
        png_bytes = png_buf.tobytes()

        doc = fitz.open()
        # A4-ish page sized to match image
        page = doc.new_page(width=w, height=h)
        page.insert_image(fitz.Rect(0, 0, w, h), stream=png_bytes)

        # Insert invisible text layer from OCR regions
        for region in (result.pages[0].regions if result.pages else []):
            bb = region.bounding_box  # [x0, y0, x1, y1]
            rect = fitz.Rect(bb[0], bb[1], bb[2], bb[3])
            # invisible text: white color, tiny font — still searchable
            page.insert_textbox(rect, region.text, fontsize=8,
                                 color=(1, 1, 1), overlay=False)

        pdf_bytes = doc.tobytes()
        doc.close()
        words = len(result.pages[0].words) if result.pages else 0
        _record("/ocr/export/searchable-pdf", words=words, ms=ms, success=True)
        fname = (file.filename or "output").rsplit(".", 1)[0]
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fname}_searchable.pdf"'},
        )
    except Exception as e:
        _record("/ocr/export/searchable-pdf", success=False, error=str(e))
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {e}")


# ─────────────────────────────────────────────
# 2. EXCEL EXPORT
# ─────────────────────────────────────────────
@router.post("/export/excel")
async def export_excel(
    file: UploadFile = File(...),
    image_service: ImageOCRService = Depends(get_image_service),
    settings: Settings = Depends(get_settings),
) -> Response:
    """Run OCR and export regions as an Excel workbook."""
    request_id = generate_request_id()
    validator = FileValidator()
    content = await validator.validate_upload(file, settings)

    t0 = time.monotonic()
    ext = Path(file.filename or "").suffix.lower()

    if ext == ".pdf":
        from api.dependencies import get_pdf_service
        pdf_svc = get_pdf_service()
        result = pdf_svc.process(content, file.filename or "doc.pdf", request_id)
    else:
        result = image_service.process(content, file.filename or "image", request_id)

    ms = (time.monotonic() - t0) * 1000

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        # Summary sheet
        ws_sum = wb.active
        ws_sum.title = "Summary"
        ws_sum.append(["File", result.filename])
        ws_sum.append(["Pages", len(result.pages)])
        ws_sum.append(["Processing Time (ms)", round(result.processing_time_ms, 1)])
        ws_sum.append(["Total Words", sum(len(p.words) for p in result.pages)])

        # One sheet per page
        for page in result.pages:
            ws = wb.create_sheet(title=f"Page {page.page_number}")
            header = ["Region #", "Text", "Confidence", "X_min", "Y_min", "X_max", "Y_max"]
            ws.append(header)
            # Style header row
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="6366F1")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")

            for i, r in enumerate(page.regions, 1):
                bb = r.bounding_box
                ws.append([i, r.text, round(r.confidence, 4), bb[0], bb[1], bb[2], bb[3]])

            # Full text sheet
            ws_txt = wb.create_sheet(title=f"Text P{page.page_number}")
            ws_txt.append(["Full Text"])
            ws_txt["A1"].font = Font(bold=True)
            for line in page.full_text.splitlines():
                ws_txt.append([line])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        words = sum(len(p.words) for p in result.pages)
        _record("/ocr/export/excel", words=words, ms=ms, success=True)
        fname = (file.filename or "output").rsplit(".", 1)[0]
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}_ocr.xlsx"'},
        )
    except Exception as e:
        _record("/ocr/export/excel", success=False, error=str(e))
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {e}")


# ─────────────────────────────────────────────
# 3. IMAGE PREPROCESSING PIPELINE (expose as endpoint)
# ─────────────────────────────────────────────
class PreprocessResponse(BaseModel):
    request_id: str
    original_size: list[int]
    processed_size: list[int]
    steps_applied: list[str]
    processing_time_ms: float


@router.post("/preprocess")
async def preprocess_image(
    file: UploadFile = File(...),
    grayscale: bool = Query(default=True),
    denoise: bool = Query(default=True),
    enhance_contrast: bool = Query(default=True),
    threshold: bool = Query(default=False),
    deskew: bool = Query(default=True),
    upscale: bool = Query(default=True),
    return_image: bool = Query(default=True, description="Return processed image bytes"),
    settings: Settings = Depends(get_settings),
) -> Response:
    """Apply configurable preprocessing steps and return the processed image."""
    content = await file.read()
    t0 = time.monotonic()

    arr = np.frombuffer(content, np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        raise HTTPException(status_code=400, detail="Cannot decode image")

    orig_h, orig_w = img.shape[:2]
    steps = []

    if deskew:
        from preprocessing.preprocessor import ImagePreprocessor
        pp = ImagePreprocessor(settings)
        img = pp._auto_rotate(img)
        steps.append("deskew")

    if grayscale:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if len(img.shape) == 3 else img
        steps.append("grayscale")

    if denoise and len(img.shape) == 2:
        img = cv2.fastNlMeansDenoising(img, h=10)
        steps.append("denoise")

    if enhance_contrast and len(img.shape) == 2:
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        img = clahe.apply(img)
        steps.append("enhance_contrast")

    if threshold and len(img.shape) == 2:
        img = cv2.adaptiveThreshold(img, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
        steps.append("adaptive_threshold")

    if upscale:
        h, w = img.shape[:2]
        if max(h, w) < 1000:
            scale = 1000 / max(h, w)
            img = cv2.resize(img, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_CUBIC)
            steps.append(f"upscale_x{scale:.1f}")

    proc_h, proc_w = img.shape[:2]
    ms = (time.monotonic() - t0) * 1000

    if return_image:
        _, buf = cv2.imencode(".png", img)
        return Response(
            content=buf.tobytes(),
            media_type="image/png",
            headers={
                "X-Steps": ",".join(steps),
                "X-Original-Size": f"{orig_w}x{orig_h}",
                "X-Processed-Size": f"{proc_w}x{proc_h}",
                "X-Processing-Ms": str(round(ms, 2)),
            },
        )

    return Response(
        content=json.dumps({
            "request_id": generate_request_id(),
            "original_size": [orig_w, orig_h],
            "processed_size": [proc_w, proc_h],
            "steps_applied": steps,
            "processing_time_ms": round(ms, 2),
        }),
        media_type="application/json",
    )


# ─────────────────────────────────────────────
# 4. BARCODE / QR DETECTION
# ─────────────────────────────────────────────
class BarcodeResult(BaseModel):
    request_id: str
    filename: str
    barcodes_found: int
    processing_time_ms: float
    barcodes: list[dict]


@router.post("/barcode", response_model=BarcodeResult)
async def detect_barcode(file: UploadFile = File(...)) -> BarcodeResult:
    """Detect barcodes and QR codes in an image."""
    request_id = generate_request_id()
    content = await file.read()
    t0 = time.monotonic()

    arr = np.frombuffer(content, np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        raise HTTPException(status_code=400, detail="Cannot decode image")

    results = []
    try:
        from pyzbar.pyzbar import decode as zbar_decode
        decoded = zbar_decode(img)
        for d in decoded:
            pts = d.polygon
            bbox = [int(d.rect.left), int(d.rect.top),
                    int(d.rect.left + d.rect.width), int(d.rect.top + d.rect.height)]
            results.append({
                "type": d.type,
                "data": d.data.decode("utf-8", errors="replace"),
                "bbox": bbox,
                "quality": d.quality,
            })
    except ImportError:
        # Fallback: OpenCV QR detector
        qr = cv2.QRCodeDetector()
        data, pts, _ = qr.detectAndDecode(img)
        if data:
            results.append({"type": "QR_CODE", "data": data, "bbox": [], "quality": 0})

    ms = (time.monotonic() - t0) * 1000
    _record("/ocr/barcode", ms=ms, success=True)
    return BarcodeResult(
        request_id=request_id,
        filename=file.filename or "image",
        barcodes_found=len(results),
        processing_time_ms=round(ms, 2),
        barcodes=results,
    )


# ─────────────────────────────────────────────
# 5. DOCUMENT CLASSIFICATION
# ─────────────────────────────────────────────
_DOC_PATTERNS = {
    "invoice":      ["invoice", "bill to", "amount due", "subtotal", "tax", "total", "payment"],
    "receipt":      ["receipt", "thank you", "change", "cashier", "store", "item", "qty"],
    "resume":       ["experience", "education", "skills", "objective", "references", "curriculum vitae", "cv"],
    "passport":     ["passport", "nationality", "date of birth", "place of birth", "expiry", "mrz"],
    "id_card":      ["national id", "id card", "cnic", "identity card", "nic", "identification"],
    "bank_statement": ["account number", "statement", "debit", "credit", "balance", "transaction", "bank"],
    "medical":      ["patient", "diagnosis", "prescription", "doctor", "hospital", "clinic", "medicine"],
    "newspaper":    ["published", "editor", "reporter", "headline", "article", "journalist"],
    "research":     ["abstract", "introduction", "methodology", "conclusion", "references", "doi", "journal"],
    "form":         ["please fill", "applicant", "signature", "date", "full name", "address", "form"],
}

class DocClassResult(BaseModel):
    request_id: str
    filename: str
    predicted_class: str
    confidence: float
    all_scores: dict[str, float]
    processing_time_ms: float
    word_count: int


@router.post("/classify", response_model=DocClassResult)
async def classify_document(
    file: UploadFile = File(...),
    image_service: ImageOCRService = Depends(get_image_service),
    settings: Settings = Depends(get_settings),
) -> DocClassResult:
    """Classify document type using OCR text + keyword heuristics."""
    request_id = generate_request_id()
    validator = FileValidator()
    content = await validator.validate_upload(file, settings)
    t0 = time.monotonic()

    result = image_service.process(content, file.filename or "doc", request_id)
    full_text = " ".join(p.full_text for p in result.pages).lower()
    word_count = sum(len(p.words) for p in result.pages)

    scores: dict[str, float] = {}
    for doc_type, keywords in _DOC_PATTERNS.items():
        hits = sum(1 for kw in keywords if kw in full_text)
        scores[doc_type] = round(hits / len(keywords), 4)

    best = max(scores, key=scores.get) if scores else "unknown"
    best_score = scores.get(best, 0.0)
    if best_score == 0.0:
        best = "general"

    ms = (time.monotonic() - t0) * 1000
    _record("/ocr/classify", words=word_count, ms=ms, success=True)
    return DocClassResult(
        request_id=request_id,
        filename=file.filename or "doc",
        predicted_class=best,
        confidence=best_score,
        all_scores=scores,
        processing_time_ms=round(ms, 2),
        word_count=word_count,
    )


# ─────────────────────────────────────────────
# 6. LAYOUT VISUALIZATION
# ─────────────────────────────────────────────
@router.post("/visualize")
async def visualize_layout(
    file: UploadFile = File(...),
    image_service: ImageOCRService = Depends(get_image_service),
    settings: Settings = Depends(get_settings),
    show_confidence: bool = Query(default=True),
    color_by_confidence: bool = Query(default=True),
) -> Response:
    """Return annotated image with OCR bounding boxes drawn."""
    request_id = generate_request_id()
    validator = FileValidator()
    content = await validator.validate_upload(file, settings)

    result = image_service.process(content, file.filename or "image", request_id)

    arr = np.frombuffer(content, np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        raise HTTPException(status_code=400, detail="Cannot decode image")

    h, w = img.shape[:2]
    if max(h, w) < 1000:
        scale = 1000 / max(h, w)
        img = cv2.resize(img, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_CUBIC)

    regions = result.pages[0].regions if result.pages else []
    for region in regions:
        bb = region.bounding_box
        conf = region.confidence
        if color_by_confidence:
            if conf > 0.8:
                color = (34, 197, 94)    # green
            elif conf > 0.5:
                color = (251, 191, 36)   # yellow
            else:
                color = (239, 68, 68)    # red
        else:
            color = (99, 102, 241)       # accent purple

        cv2.rectangle(img, (bb[0], bb[1]), (bb[2], bb[3]), color, 2)
        if show_confidence:
            label = f"{conf*100:.0f}%"
            cv2.putText(img, label, (bb[0], max(bb[1]-4, 10)),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.45, color, 1, cv2.LINE_AA)

    _, buf = cv2.imencode(".jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 90])
    return Response(
        content=buf.tobytes(),
        media_type="image/jpeg",
        headers={
            "X-Regions-Count": str(len(regions)),
            "X-Request-ID": request_id,
        },
    )


# ─────────────────────────────────────────────
# 7. BACKGROUND JOB QUEUE
# ─────────────────────────────────────────────
class JobStatus(BaseModel):
    job_id: str
    status: str          # pending | running | done | failed
    filename: str
    created_at: float
    completed_at: Optional[float] = None
    processing_time_ms: Optional[float] = None
    word_count: Optional[int] = None
    error: Optional[str] = None
    result_preview: Optional[str] = None


@router.post("/jobs/submit")
async def submit_job(
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = None,
) -> dict:
    """Submit a large file for background OCR processing."""
    job_id = str(uuid.uuid4())
    content = await file.read()
    filename = file.filename or "file"
    _jobs[job_id] = {
        "job_id": job_id, "status": "pending",
        "filename": filename, "created_at": time.time(),
        "completed_at": None, "processing_time_ms": None,
        "word_count": None, "error": None, "result_preview": None,
    }
    background_tasks.add_task(_process_job, job_id, content, filename)
    return {"job_id": job_id, "status": "pending", "message": f"Job {job_id} queued"}


async def _process_job(job_id: str, content: bytes, filename: str):
    """Background task: run OCR and store result."""
    from api.dependencies import get_image_service, get_settings
    _jobs[job_id]["status"] = "running"
    t0 = time.monotonic()
    try:
        svc = get_image_service()
        ext = Path(filename).suffix.lower()
        if ext == ".pdf":
            from api.dependencies import get_pdf_service
            svc = get_pdf_service()
            result = svc.process(content, filename, job_id)
        else:
            result = svc.process(content, filename, job_id)
        ms = (time.monotonic() - t0) * 1000
        words = sum(len(p.words) for p in result.pages)
        preview = result.pages[0].full_text[:200] if result.pages else ""
        _jobs[job_id].update({
            "status": "done", "completed_at": time.time(),
            "processing_time_ms": round(ms, 2), "word_count": words,
            "result_preview": preview,
        })
    except Exception as e:
        _jobs[job_id].update({"status": "failed", "completed_at": time.time(), "error": str(e)})


@router.get("/jobs/{job_id}", response_model=JobStatus)
async def get_job_status(job_id: str) -> JobStatus:
    """Get the status of a background OCR job."""
    job = _jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")
    return JobStatus(**job)


@router.get("/jobs")
async def list_jobs(limit: int = Query(default=20, ge=1, le=100)) -> dict:
    """List recent background jobs."""
    jobs = sorted(_jobs.values(), key=lambda j: j["created_at"], reverse=True)[:limit]
    return {"total": len(_jobs), "jobs": jobs}


# ─────────────────────────────────────────────
# 8. MONITORING & METRICS
# ─────────────────────────────────────────────
@router.get("/metrics")
async def get_metrics() -> dict:
    """Return service-level metrics for monitoring."""
    total = _metrics["total_requests"]
    success_rate = (_metrics["successful"] / total * 100) if total > 0 else 0.0
    avg_ms = (_metrics["total_processing_ms"] / _metrics["successful"]) if _metrics["successful"] > 0 else 0.0
    avg_wpm = (_metrics["total_words"] / _metrics["successful"]) if _metrics["successful"] > 0 else 0.0
    jobs_by_status = defaultdict(int)
    for j in _jobs.values():
        jobs_by_status[j["status"]] += 1
    return {
        "total_requests": total,
        "successful": _metrics["successful"],
        "failed": _metrics["failed"],
        "success_rate_pct": round(success_rate, 2),
        "avg_processing_ms": round(avg_ms, 2),
        "avg_words_per_request": round(avg_wpm, 1),
        "total_words_extracted": _metrics["total_words"],
        "by_endpoint": dict(_metrics["by_endpoint"]),
        "recent_errors": _metrics["recent_errors"][:5],
        "jobs": dict(jobs_by_status),
        "uptime_note": "metrics reset on restart",
    }


def record_metrics(endpoint: str, words: int = 0, ms: float = 0.0, success: bool = True, error: str = ""):
    """Public helper for other routers to record metrics."""
    _record(endpoint, words=words, ms=ms, success=success, error=error)
