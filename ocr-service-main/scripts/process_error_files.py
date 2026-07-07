#!/usr/bin/env python3
"""
Batch OCR processor for error files folder.
Uses the same MistralOCREngine + PaddleOCREngine (fallback) as test_pdf.py.

- PDFs  → MistralOCREngine (primary) → PaddleOCREngine (fallback)
- Excel → openpyxl direct parse (no OCR needed)

Results saved to:
  output/error_files_results/results.json       ← full data per file
  output/error_files_results/summary.csv        ← one row per file
  output/error_files_results/texts/<file>.txt   ← extracted text per file

Usage:
    python3 process_error_files.py
    python3 process_error_files.py --input /path/to/files --output /path/to/output
    python3 process_error_files.py --file "somefile.pdf"
    python3 process_error_files.py --skip-excel
    python3 process_error_files.py --skip-pdf
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
import traceback
from pathlib import Path

# ── Project root on sys.path ─────────────────────────────────────────────────
# Script lives in scripts/ — go one level up to reach project root
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from config.settings import get_settings
from ocr.engines.mistral_engine import MistralOCREngine
from ocr.engines.paddle_engine import PaddleOCREngine

# ── Defaults ─────────────────────────────────────────────────────────────────
DEFAULT_INPUT_DIR  = Path("/home/tenbitsolutions/Downloads/ocr-testing/error files")
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "output" / "error_files_results"

SUPPORTED_PDF_EXTS   = frozenset({".pdf"})
SUPPORTED_EXCEL_EXTS = frozenset({".xlsx", ".xls"})
SUPPORTED_PPT_EXTS   = frozenset({".pptx", ".ppt"})
SUPPORTED_EXTS       = SUPPORTED_PDF_EXTS | SUPPORTED_EXCEL_EXTS | SUPPORTED_PPT_EXTS


# ─────────────────────────────────────────────────────────────────────────────
# Engine initialisation (once, shared across all files)
# ─────────────────────────────────────────────────────────────────────────────

def init_engines():
    settings = get_settings()
    print("  Initialising OCR engines...")

    mistral = MistralOCREngine(api_key=settings.mistral_api_key)
    if mistral.is_available():
        print("  ✓ MistralOCREngine  — ready (primary)")
    else:
        print("  ✗ MistralOCREngine  — unavailable (no API key?)")
        mistral = None

    langs = [l.strip() for l in settings.ocr_languages.split(",")]
    try:
        paddle = PaddleOCREngine(languages=langs, use_gpu=settings.use_gpu)
        if paddle.is_available():
            print("  ✓ PaddleOCREngine   — ready (fallback)")
        else:
            print("  ✗ PaddleOCREngine   — unavailable")
            paddle = None
    except Exception as e:
        print(f"  ✗ PaddleOCREngine   — init error: {e}")
        paddle = None

    return mistral, paddle


# ─────────────────────────────────────────────────────────────────────────────
# PDF processing
# ─────────────────────────────────────────────────────────────────────────────

def process_pdf_bytes(
    pdf_bytes: bytes,
    filename: str,
    mistral: MistralOCREngine | None,
    paddle: PaddleOCREngine | None,
) -> dict:
    """Run OCR on raw PDF bytes. Mistral first, Paddle fallback. Returns result dict."""
    engine_used = None
    pages = []

    # ── Try Mistral ───────────────────────────────────────────────────────────
    if mistral:
        try:
            print(f"    [Mistral] Processing {filename}...")
            pages = mistral.process_pdf(pdf_bytes, filename)
            engine_used = "mistral"
            print(f"    [Mistral] ✓ {len(pages)} page(s)")
        except Exception as e:
            print(f"    [Mistral] ✗ Failed: {e}")
            print(f"    [Paddle]  Falling back...")
            pages = []

    # ── Fallback: Paddle ──────────────────────────────────────────────────────
    if not pages and paddle:
        try:
            print(f"    [Paddle]  Processing {filename}...")
            pages = paddle.process_pdf(pdf_bytes, filename)
            engine_used = "paddle"
            print(f"    [Paddle]  ✓ {len(pages)} page(s)")
        except Exception as e:
            raise RuntimeError(f"Both engines failed. Paddle error: {e}") from e

    if not pages:
        raise RuntimeError("No OCR engine available to process this PDF.")

    serialised_pages = [p.model_dump() for p in pages]
    full_text = "\n\n".join(p.full_text for p in pages if p.full_text)

    return {
        "engine": engine_used,
        "total_pages": len(pages),
        "pages": serialised_pages,
        "full_text": full_text,
    }


def process_pdf(file_path: Path, mistral: MistralOCREngine | None, paddle: PaddleOCREngine | None) -> dict:
    """Run OCR on a PDF file."""
    return process_pdf_bytes(file_path.read_bytes(), file_path.name, mistral, paddle)


# ─────────────────────────────────────────────────────────────────────────────
# PPT processing  (LibreOffice → PDF → Mistral, fallback Paddle)
# ─────────────────────────────────────────────────────────────────────────────

def process_ppt(
    file_path: Path,
    mistral: MistralOCREngine | None,
    paddle: PaddleOCREngine | None,
) -> dict:
    """Convert PPT/PPTX to PDF via LibreOffice, then run OCR (Mistral → Paddle fallback)."""
    import subprocess
    import tempfile
    import shutil

    filename = file_path.name
    print(f"    [LibreOffice] Converting {filename} to PDF...")

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)

        # Copy the file into the temp dir (LibreOffice needs a clean path)
        src_copy = tmp_path / file_path.name
        shutil.copy2(file_path, src_copy)

        # Convert to PDF using LibreOffice headless
        result = subprocess.run(
            [
                "libreoffice", "--headless", "--convert-to", "pdf",
                "--outdir", str(tmp_path), str(src_copy),
            ],
            capture_output=True, text=True, timeout=120,
        )

        if result.returncode != 0:
            raise RuntimeError(
                f"LibreOffice conversion failed: {result.stderr.strip() or result.stdout.strip()}"
            )

        # Find the generated PDF
        pdf_files = list(tmp_path.glob("*.pdf"))
        if not pdf_files:
            raise RuntimeError("LibreOffice did not produce a PDF output.")

        pdf_path = pdf_files[0]
        pdf_bytes = pdf_path.read_bytes()
        pdf_size_kb = len(pdf_bytes) // 1024
        print(f"    [LibreOffice] ✓ Converted → {pdf_path.name} ({pdf_size_kb} KB)")

    # Now OCR the PDF: Mistral primary, Paddle fallback
    ocr_data = process_pdf_bytes(pdf_bytes, filename, mistral, paddle)
    ocr_data["converted_via"] = "libreoffice"
    return ocr_data


# ─────────────────────────────────────────────────────────────────────────────
# Excel processing
# ─────────────────────────────────────────────────────────────────────────────

def process_excel(file_path: Path) -> dict:
    """Extract all sheets and cell data from an Excel file via openpyxl."""
    import openpyxl

    print(f"    [Excel]  Parsing {file_path.name}...")
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    sheets = []
    all_text_parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_data = []
        sheet_lines = []

        for row in ws.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue
            row_values = [str(c) if c is not None else "" for c in row]
            rows_data.append(row_values)
            line = " | ".join(v for v in row_values if v.strip())
            if line.strip():
                sheet_lines.append(line)

        sheet_text = "\n".join(sheet_lines)
        all_text_parts.append(f"=== Sheet: {sheet_name} ===\n{sheet_text}")

        sheets.append({
            "sheet_name": sheet_name,
            "row_count": len(rows_data),
            "column_count": ws.max_column or 0,
            "text": sheet_text,
            "rows": rows_data[:500],   # cap to keep JSON manageable
        })

    wb.close()
    full_text = "\n\n".join(all_text_parts)
    print(f"    [Excel]  ✓ {len(sheets)} sheet(s), {len(full_text.split())} words")

    return {
        "engine": "openpyxl",
        "sheet_count": len(sheets),
        "total_pages": len(sheets),
        "sheets": sheets,
        "full_text": full_text,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Per-file dispatcher
# ─────────────────────────────────────────────────────────────────────────────

def process_file(
    file_path: Path,
    mistral: MistralOCREngine | None,
    paddle: PaddleOCREngine | None,
) -> dict:
    ext = file_path.suffix.lower()
    start = time.monotonic()

    result: dict = {
        "filename":           file_path.name,
        "filepath":           str(file_path),
        "extension":          ext,
        "file_size_kb":       round(file_path.stat().st_size / 1024, 2),
        "status":             "success",
        "engine_used":        None,
        "processing_time_ms": 0,
        "error":              None,
        "data":               None,
    }

    try:
        if ext in SUPPORTED_PDF_EXTS:
            data = process_pdf(file_path, mistral, paddle)
        elif ext in SUPPORTED_PPT_EXTS:
            data = process_ppt(file_path, mistral, paddle)
        elif ext in SUPPORTED_EXCEL_EXTS:
            data = process_excel(file_path)
        else:
            raise ValueError(f"Unsupported file type: {ext!r}")

        result["data"]        = data
        result["engine_used"] = data.get("engine", "unknown")

    except Exception as e:
        result["status"] = "error"
        result["error"]  = str(e)
        print(f"    [ERROR]  {e}")

    result["processing_time_ms"] = round((time.monotonic() - start) * 1000, 1)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Save results
# ─────────────────────────────────────────────────────────────────────────────

def save_results(results: list[dict], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    texts_dir = output_dir / "texts"
    texts_dir.mkdir(exist_ok=True)

    # Full JSON
    json_path = output_dir / "results.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\n  ✓ Full JSON  → {json_path}")

    # Per-file .txt
    for r in results:
        safe = r["filename"].replace("/", "_").replace("\\", "_")
        txt_path = texts_dir / (safe + ".txt")
        if r["data"] and "full_text" in r["data"]:
            txt_path.write_text(r["data"]["full_text"], encoding="utf-8")
        elif r["status"] == "error":
            txt_path.write_text(f"ERROR: {r['error']}", encoding="utf-8")
        else:
            txt_path.write_text("", encoding="utf-8")
    print(f"  ✓ Text files → {texts_dir}/  ({len(results)} files)")

    # CSV summary
    csv_path = output_dir / "summary.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "filename", "extension", "file_size_kb", "status",
            "engine_used", "processing_time_ms",
            "total_pages_or_sheets", "total_words", "total_chars", "error",
        ])
        writer.writeheader()
        for r in results:
            ft = (r["data"] or {}).get("full_text", "") if r["data"] else ""
            writer.writerow({
                "filename":              r["filename"],
                "extension":             r["extension"],
                "file_size_kb":          r["file_size_kb"],
                "status":                r["status"],
                "engine_used":           r.get("engine_used") or "",
                "processing_time_ms":    r["processing_time_ms"],
                "total_pages_or_sheets": (r["data"] or {}).get("total_pages", 0),
                "total_words":           len(ft.split()),
                "total_chars":           len(ft),
                "error":                 r.get("error") or "",
            })
    print(f"  ✓ Summary CSV → {csv_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Console summary
# ─────────────────────────────────────────────────────────────────────────────

def print_summary(results: list[dict], elapsed: float) -> None:
    ok  = [r for r in results if r["status"] == "success"]
    err = [r for r in results if r["status"] == "error"]

    print("\n" + "═" * 72)
    print("  BATCH COMPLETE")
    print("═" * 72)
    print(f"  Total    : {len(results)}  |  ✓ Success: {len(ok)}  |  ✗ Failed: {len(err)}  |  ⏱  {elapsed:.1f}s")
    print("─" * 72)

    for r in results:
        icon = "✓" if r["status"] == "success" else "✗"
        d    = r.get("data") or {}
        ft   = d.get("full_text", "")
        pages = d.get("total_pages", 0)
        words = len(ft.split())
        eng   = r.get("engine_used") or "—"
        extra = f"{pages}p  {words:,}w" if r["status"] == "success" else (r.get("error") or "")[:55]
        print(f"  {icon}  {r['filename']:<52}  [{eng:<20}]  {extra}")

    if err:
        print("\n  Failed files:")
        for r in err:
            print(f"    ✗  {r['filename']}: {r.get('error','')}")
    print("═" * 72)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Batch OCR — error files folder")
    parser.add_argument("--input",       type=Path, default=DEFAULT_INPUT_DIR)
    parser.add_argument("--output",      type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--file",        type=str,  default=None,
                        help="Process a single file by name")
    parser.add_argument("--skip-excel",  action="store_true")
    parser.add_argument("--skip-pdf",    action="store_true")
    parser.add_argument("--skip-ppt",    action="store_true")
    args = parser.parse_args()

    if not args.input.exists():
        print(f"ERROR: Input directory not found: {args.input}")
        sys.exit(1)

    # Init engines once
    print(f"\n{'═'*72}")
    print("  OCR Batch Processor — error files")
    print(f"{'═'*72}")
    mistral, paddle = init_engines()
    print()

    # Collect files
    if args.file:
        fp = args.input / args.file
        if not fp.exists():
            print(f"ERROR: File not found: {fp}")
            sys.exit(1)
        all_files = [fp]
    else:
        all_files = sorted(
            p for p in args.input.iterdir()
            if p.is_file() and p.suffix.lower() in SUPPORTED_EXTS
        )

    if args.skip_excel:
        all_files = [f for f in all_files if f.suffix.lower() not in SUPPORTED_EXCEL_EXTS]
    if args.skip_pdf:
        all_files = [f for f in all_files if f.suffix.lower() not in SUPPORTED_PDF_EXTS]
    if args.skip_ppt:
        all_files = [f for f in all_files if f.suffix.lower() not in SUPPORTED_PPT_EXTS]

    n_pdf   = sum(1 for f in all_files if f.suffix.lower() in SUPPORTED_PDF_EXTS)
    n_excel = sum(1 for f in all_files if f.suffix.lower() in SUPPORTED_EXCEL_EXTS)
    n_ppt   = sum(1 for f in all_files if f.suffix.lower() in SUPPORTED_PPT_EXTS)

    print(f"  Input  : {args.input}")
    print(f"  Output : {args.output}")
    print(f"  Files  : {len(all_files)} total  ({n_pdf} PDF, {n_excel} Excel, {n_ppt} PPT/PPTX)")
    print(f"{'═'*72}\n")

    if not all_files:
        print("No supported files found.")
        sys.exit(0)

    results = []
    t0 = time.monotonic()

    for i, fp in enumerate(all_files, 1):
        print(f"[{i}/{len(all_files)}] {fp.name}  ({fp.stat().st_size//1024} KB)")
        r = process_file(fp, mistral, paddle)
        results.append(r)
        icon = "✓" if r["status"] == "success" else "✗"
        print(f"  {icon} {r['processing_time_ms']:.0f}ms  engine={r['engine_used'] or 'N/A'}\n")

    elapsed = time.monotonic() - t0
    save_results(results, args.output)
    print_summary(results, elapsed)


if __name__ == "__main__":
    main()
