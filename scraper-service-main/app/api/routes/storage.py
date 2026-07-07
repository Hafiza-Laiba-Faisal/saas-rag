"""
Storage routes — sessions, posts, stats, excel export.
No DB code here — delegates to storage layer.
"""
from __future__ import annotations
import io
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from storage.sqlite_storage import default_storage
from config.settings import EXCEL_IMG_WIDTH, EXCEL_IMG_HEIGHT, EXCEL_ROW_HEIGHT, EXCEL_HEADER_COLOR

router = APIRouter(prefix="/db", tags=["storage"])


# ── Sessions ──────────────────────────────────────────────────────────────────

@router.get("/sessions")
def list_sessions():
    return {"sessions": default_storage.list_sessions()}


@router.get("/sessions/{session_id}")
def get_session(session_id: int):
    s = default_storage.get_session(session_id)
    if not s:
        raise HTTPException(404, "Session not found")
    return s


@router.delete("/sessions/{session_id}")
def delete_session(session_id: int):
    if not default_storage.delete_session(session_id):
        raise HTTPException(404, "Session not found")
    return {"deleted": True, "session_id": session_id}


# ── Posts ─────────────────────────────────────────────────────────────────────

@router.get("/posts")
def get_posts(
    search:       str = Query(""),
    page_url:     str = Query(""),
    content_type: str = Query(""),
    date_from:    str = Query(""),
    date_to:      str = Query(""),
    limit:        int = Query(50, ge=1, le=200),
    offset:       int = Query(0, ge=0),
):
    return default_storage.get_all_posts(
        search=search, page_url_filter=page_url, content_type=content_type,
        date_from=date_from, date_to=date_to, limit=limit, offset=offset,
    )


@router.delete("/posts/{post_db_id}")
def delete_post(post_db_id: int):
    if not default_storage.delete_post(post_db_id):
        raise HTTPException(404, "Post not found")
    return {"deleted": True, "id": post_db_id}


@router.get("/stats")
def get_stats():
    return default_storage.get_stats()


# ── Excel export ──────────────────────────────────────────────────────────────

@router.get("/export/excel")
def export_excel(
    page_url:     str = Query(""),
    content_type: str = Query("post"),
    date_from:    str = Query(""),
    date_to:      str = Query(""),
    post_ids:     str = Query(""),
    limit:        int = Query(500, ge=1, le=2000),
):
    import requests as req_lib
    import openpyxl, openpyxl.drawing.image
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from utils.media import resize_image

    if post_ids.strip():
        id_list  = [p.strip() for p in post_ids.split(",") if p.strip()]
        data     = default_storage.get_all_posts(content_type=content_type, limit=len(id_list)+10)
        id_order = {pid: i for i, pid in enumerate(id_list)}
        id_set   = set(id_list)
        posts    = sorted(
            [p for p in data.get("posts", []) if p.get("post_id","") in id_set],
            key=lambda p: id_order.get(p.get("post_id",""), 999),
        )
    else:
        data  = default_storage.get_all_posts(
            page_url_filter=page_url, content_type=content_type,
            date_from=date_from, date_to=date_to, limit=limit,
        )
        posts = data.get("posts", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reels" if content_type == "reel" else "Posts"

    hfill  = PatternFill("solid", fgColor=EXCEL_HEADER_COLOR)
    hfont  = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tleft  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(["#", "Image", "Caption", "Date", "Post URL"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 22
    for col, w in zip("ABCDE", [5, 25, 60, 18, 45]):
        ws.column_dimensions[col].width = w

    sess = req_lib.Session()
    sess.headers.update({"User-Agent": "Mozilla/5.0 Chrome/124", "Referer": "https://www.facebook.com/"})

    for row_idx, post in enumerate(posts, 2):
        ws.row_dimensions[row_idx].height = EXCEL_ROW_HEIGHT
        media   = post.get("media") or []
        caption = (post.get("caption") or "").strip()
        date    = (post.get("posted_at") or "")[:16].replace("T", " ")
        url     = post.get("post_url") or ""
        img_url = next((m["url"] for m in media if m.get("type") == "image" and m.get("url")), "")
        if not img_url:
            img_url = next((m.get("thumb") or m.get("url","") for m in media if m.get("type") == "video"), "")

        c = ws.cell(row=row_idx, column=1, value=row_idx-1); c.alignment=center; c.border=border
        ws.cell(row=row_idx, column=2).border=border; ws.cell(row=row_idx, column=2).alignment=center
        if img_url:
            try:
                r = sess.get(img_url, timeout=10, stream=True)
                buf = resize_image(r.content, EXCEL_IMG_WIDTH, EXCEL_IMG_HEIGHT)
                if buf:
                    xl_img = openpyxl.drawing.image.Image(buf)
                    xl_img.width = EXCEL_IMG_WIDTH; xl_img.height = EXCEL_IMG_HEIGHT
                    xl_img.anchor = f"B{row_idx}"; ws.add_image(xl_img)
                else:
                    ws.cell(row=row_idx, column=2, value="—")
            except Exception:
                ws.cell(row=row_idx, column=2, value="—")

        c = ws.cell(row=row_idx, column=3, value=caption or "—"); c.alignment=tleft; c.border=border
        c = ws.cell(row=row_idx, column=4, value=date or "—"); c.alignment=center; c.border=border
        c = ws.cell(row=row_idx, column=5, value=url or "—")
        if url:
            c.hyperlink = url; c.font = Font(color="1877F2", underline="single")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); c.border=border

    ws.freeze_panes = "A2"
    out = io.BytesIO(); wb.save(out); out.seek(0)
    label = "reels" if content_type == "reel" else "posts"
    return StreamingResponse(out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="fb-{label}-{date_from or "all"}.xlsx"'})
